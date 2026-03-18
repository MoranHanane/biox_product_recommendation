from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Optional, Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows



# Validation stricte de l'index de similarité (FORMAT UNIQUE)
########################################

REQUIRED_INDEX_COLS = {
    "product_code_raw_A",
    "product_code_base_A",
    "Reference_A",
    "product_code_raw_B",
    "product_code_base_B",
    "Reference_B",
    "species_B",
    "class_name_B",
    "similarity",
}


def _validate_similarity_index(df: pd.DataFrame) -> None:
    """
    Vérifie que l’index de similarité respecte le format attendu.

    Contrôle :
    - Présence stricte des colonnes obligatoires

    Lève une erreur bloquante si le format est invalide.
    """
    missing = REQUIRED_INDEX_COLS - set(df.columns)
    if missing:
        raise ValueError(
            "Index de similarité invalide.\n"
            f"Colonnes manquantes : {sorted(missing)}\n"
            "Format attendu STRICT :\n"
            + ", ".join(sorted(REQUIRED_INDEX_COLS))
        )



# Helpers master mapping
###########################

def build_master_maps(master_df: pd.DataFrame) -> Dict[str, Dict[str, Any]]:
    """
    Mapping par Reference (libellé complet) -> infos nécessaires aux overrides/substitutions & MPC_range.
    IMPORTANT: on ne modifie PAS les espaces internes; On se contente de strip() pour robustesse sur espaces en bord.
    """
    required = {"Reference", "product_code_raw", "product_code_base", "species", "class_name", "MPC_range",
                "commentaire_produit", "customer_product_suggestions"}
    missing = required - set(master_df.columns)
    if missing:
        raise ValueError(f"master_df invalide pour overrides. Colonnes manquantes: {sorted(missing)}")

    m = {}
    for _, r in master_df.iterrows():
        ref = r["Reference"]
        if not isinstance(ref, str):
            continue
        ref_key = ref.strip()
        m[ref_key] = {
            "Reference": ref,
            "product_code_raw": str(r["product_code_raw"]) if pd.notna(r["product_code_raw"]) else None,
            "product_code_base": str(r["product_code_base"]) if pd.notna(r["product_code_base"]) else None,
            "species": r["species"] if pd.notna(r["species"]) else None,
            "class_name": r["class_name"] if pd.notna(r["class_name"]) else None,
            "MPC_range": r["MPC_range"] if pd.notna(r["MPC_range"]) else "Not applicable",
            "commentaire_produit": r["commentaire_produit"] if pd.notna(r["commentaire_produit"]) else None,
            "customer_product_suggestions": (
                r["customer_product_suggestions"].strip()
                if isinstance(r["customer_product_suggestions"], str)
                else None
            ),
        }
    return m



# Core
###############

def make_product_suggestions_for_customer(
    lines: pd.DataFrame,
    similarity_index: pd.DataFrame,
    master_df: Optional[pd.DataFrame] = None,
    allowed_species: Optional[List[str]] = None,
    allowed_classes: Optional[List[str]] = None,
    top_k: int = 5,
) -> pd.DataFrame:
    """
    V4+ (override rules + MPC_range_suggested)

    A1) override = Reference_B unique -> duplique sur top_k lignes
    A2) override = "!Aucune proposition pour ce produit!" -> aucune suggestion, no_suggestion_reason = commentaire_produit
    A3) override vide -> algo normal (reweight déjà fait côté index)

    - override IGNORE les filtres, mais on met un message dans single_suggestion_reason
    - "Not applicable" en toutes lettres quand non concerné (NB: REGLE A SUPPRIMER)
    """

    
    # 1) Validation entrées
    #################################
    required_lines_cols = {
        "product_code_raw",
        "product_code_base",
        "Reference",
        "CA_N",
        "CA_N1",
    }
    missing = required_lines_cols - set(lines.columns)
    if missing:
        raise ValueError(f"`lines` invalide. Colonnes manquantes : {sorted(missing)}")

    _validate_similarity_index(similarity_index)

    master_map = None
    if master_df is not None:
        master_map = build_master_maps(master_df)

    
    # 2) Panier client
    ################################
    basket = lines[(lines["CA_N"] > 0) | (lines["CA_N1"] > 0)].copy()
    if basket.empty:
        return pd.DataFrame()

    basket["product_code_raw"] = basket["product_code_raw"].astype(str)
    basket["Reference"] = basket["Reference"].astype(str)

    purchased_codes = basket["product_code_raw"].unique().tolist()
    purchased_set = set(purchased_codes)

    base_map = dict(zip(basket["product_code_raw"], basket["product_code_base"]))
    ref_map = dict(zip(basket["product_code_raw"], basket["Reference"]))

    
    # 3) Préparation index
    ################################
    sim = similarity_index.copy()
    sim["similarity"] = pd.to_numeric(sim["similarity"], errors="coerce").fillna(0.0)

    # exclure A → A
    sim = sim[sim["product_code_raw_A"] != sim["product_code_raw_B"]]

    # Filtres UI (côté B)
    sim_filtered = sim
    if allowed_species is not None:
        sim_filtered = sim_filtered[sim_filtered["species_B"].isin(allowed_species)]
    if allowed_classes is not None:
        sim_filtered = sim_filtered[sim_filtered["class_name_B"].isin(allowed_classes)]

    # Marquage only
    sim_filtered = sim_filtered.copy()
    sim_filtered["already_in_basket"] = sim_filtered["product_code_raw_B"].isin(purchased_set)

    # MPC_range mapping côté B (si master fourni)
    b_mpc_map = {}
    b_meta_map = {}
    if master_map is not None:
        for k, v in master_map.items():
            code_raw = v.get("product_code_raw")
            if code_raw:
                b_mpc_map[str(code_raw)] = v.get("MPC_range", "Not applicable")
                b_meta_map[str(code_raw)] = v  # pour retrouver Reference_B, species, class

    #
    # 4) Suggestions (override prioritaire, puis fallback algo quand override ne s'applique pas)
    ################################################
    rows = []

    for code_a in purchased_codes:
        base_a = base_map.get(code_a)
        ref_a = ref_map.get(code_a)

        ref_a_key = ref_a.strip() if isinstance(ref_a, str) else None

        # ---------- OVERRIDE (si master fourni) ----------
        if master_map is not None and ref_a_key in master_map:
            override = master_map[ref_a_key].get("customer_product_suggestions")
            comment = master_map[ref_a_key].get("commentaire_produit")

            # A2 : aucune proposition
            if isinstance(override, str) and override.strip() == "!Aucune proposition pour ce produit!":
                for _ in range(int(top_k)):
                    rows.append(
                        {
                            "Reference_A": ref_a,
                            "Reference_B": None,
                            "similarity": None,
                            "MPC_range_suggested": None,
                            "species_suggested": None,
                            "class_name_suggested": None,
                            "already_in_basket": None,
                            "product_code_raw_A": code_a,
                            "product_code_raw_B": None,
                            "product_code_base_A": base_a,
                            "product_code_base_B": None,
                            "no_suggestion_reason": comment if comment else "!Aucune proposition pour ce produit!",
                            "single_suggestion_reason": "Not applicable",
                        }
                    )
                continue

            # A1 : suggestion unique imposée (Reference complet)
            if isinstance(override, str) and override.strip():
                target_ref = override.strip()
                target = master_map.get(target_ref)

                # Si target_ref n'existe pas dans master -> fallback algo normal + warning
                if target is not None and target.get("product_code_raw"):
                    target_code_b = str(target["product_code_raw"])
                    target_base_b = str(target["product_code_base"]) if target.get("product_code_base") else None

                    # IMPORTANT: exclusion même base (règle globale)
                    if target_base_b is not None and base_a is not None and str(target_base_b) == str(base_a):
                        # Si l'override viole la règle base, on force "aucune suggestion"
                        for _ in range(int(top_k)):
                            rows.append(
                                {
                                    "Reference_A": ref_a,
                                    "Reference_B": None,
                                    "similarity": None,
                                    "MPC_range_suggested": None,
                                    "species_suggested": None,
                                    "class_name_suggested": None,
                                    "already_in_basket": None,
                                    "product_code_raw_A": code_a,
                                    "product_code_raw_B": None,
                                    "product_code_base_A": base_a,
                                    "product_code_base_B": None,
                                    "no_suggestion_reason": (
                                        "Override invalide: produit suggéré partage la même product_code_base"
                                    ),
                                    "single_suggestion_reason": comment if comment else "Not applicable",
                                }
                            )
                        continue

                    # Similarité si existante dans l'index (non filtré, car override ignore filtres)
                    sim_match = sim[
                        (sim["product_code_raw_A"].astype(str) == str(code_a))
                        & (sim["product_code_raw_B"].astype(str) == str(target_code_b))
                    ]
                    sim_val = float(sim_match["similarity"].iloc[0]) if not sim_match.empty else None

                    # message override ignore filtres si besoin
                    warn = []
                    if allowed_species is not None and target.get("species") not in allowed_species:
                        warn.append("species")
                    if allowed_classes is not None and target.get("class_name") not in allowed_classes:
                        warn.append("class")
                    warn_txt = ""
                    if warn:
                        warn_txt = f" | OVERRIDE: filtres ignorés ({', '.join(warn)})"

                    for _ in range(int(top_k)):
                        rows.append(
                            {
                                "Reference_A": ref_a,
                                "Reference_B": target["Reference"],
                                "similarity": sim_val,
                                "MPC_range_suggested": target.get("MPC_range", "Not applicable"),
                                "species_suggested": target.get("species"),
                                "class_name_suggested": target.get("class_name"),
                                "already_in_basket": (target_code_b in purchased_set),
                                "product_code_raw_A": code_a,
                                "product_code_raw_B": target_code_b,
                                "product_code_base_A": base_a,
                                "product_code_base_B": target_base_b,
                                "no_suggestion_reason": None,
                                "single_suggestion_reason": (comment if comment else "Not applicable") + warn_txt,
                            }
                        )
                    continue

        # ---------- ALGO NORMAL (A3) ----------
        cand = sim_filtered[sim_filtered["product_code_raw_A"].astype(str) == str(code_a)].copy()

        # Exclusion même base
        if base_a is not None:
            cand = cand[cand["product_code_base_B"].astype(str) != str(base_a)]

        if cand.empty:
            rows.append(
                {
                    "Reference_A": ref_a,
                    "Reference_B": None,
                    "similarity": None,
                    "MPC_range_suggested": None,
                    "species_suggested": None,
                    "class_name_suggested": None,
                    "already_in_basket": None,
                    "product_code_raw_A": code_a,
                    "product_code_raw_B": None,
                    "product_code_base_A": base_a,
                    "product_code_base_B": None,
                    "no_suggestion_reason": (
                        "Aucune suggestion après filtres (product_code_base / species / class)"
                    ),
                    "single_suggestion_reason": "Not applicable",
                }
            )
            continue

        cand = cand.sort_values("similarity", ascending=False).head(int(top_k))

        for _, r in cand.iterrows():
            code_b = str(r["product_code_raw_B"])
            rows.append(
                {
                    "Reference_A": ref_a,
                    "Reference_B": r["Reference_B"],
                    "similarity": r["similarity"],
                    "MPC_range_suggested": b_mpc_map.get(code_b, "Not applicable") if b_mpc_map else "Not applicable",
                    "species_suggested": r["species_B"],
                    "class_name_suggested": r["class_name_B"],
                    "already_in_basket": bool(r.get("already_in_basket", False)),
                    "product_code_raw_A": code_a,
                    "product_code_raw_B": code_b,
                    "product_code_base_A": base_a,
                    "product_code_base_B": r["product_code_base_B"],
                    "no_suggestion_reason": None,
                    "single_suggestion_reason": "Not applicable",
                }
            )

    out = pd.DataFrame(rows)

    final_cols = [
        "Reference_A",
        "Reference_B",
        "similarity",
        "MPC_range_suggested",          # <-- juste après similarity
        "species_suggested",
        "class_name_suggested",
        "already_in_basket",
        "product_code_raw_A",
        "product_code_raw_B",
        "product_code_base_A",
        "product_code_base_B",
        "no_suggestion_reason",
        "single_suggestion_reason",     # <-- aux côtés de no_suggestion_reason
    ]

    return out[final_cols]



# Export Excel + mise en forme
#############################

def export_suggestions_to_excel(df: pd.DataFrame, output_path: Path | str) -> None:
    """
    Exporte les recommandations dans un fichier Excel formaté.

    Mise en forme :
    - En-têtes en gras
    - Lignes déjà présentes dans le panier en gras
    - Colonnes techniques masquées (codes produits)

    Objectif :
    - Fournir un livrable directement exploitable par les équipes métier.
    """
    output_path = Path(output_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "customer_product_suggestions"

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # En-tête en gras
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

   

    headers = [cell.value for cell in ws[1]]
    if "already_in_basket" in headers:
        idx_flag = headers.index("already_in_basket") + 1
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=idx_flag).value is True:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).font = Font(bold=True)

    # Masquer colonnes codes
    to_hide = {"product_code_raw_A", "product_code_raw_B", "product_code_base_A", "product_code_base_B"}
    for h in to_hide:
        if h in headers:
            col_idx = headers.index(h) + 1
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].hidden = True

    wb.save(output_path)